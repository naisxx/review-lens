"""
Rebuild src/data/cube.json from the raw Home Depot review corpora (xlsx).

Collapses ~440k raw reviews into the compact OLAP cube the dashboard reads:
grain (brand x subcategory x region x month), each cell carrying the star
histogram, verified/recommended/response counts, unverified + unverified-
complaint counts, and per-source-channel review counts + star sums.

Reverse-engineered to reproduce the shipped accessory cube exactly, then
extended to faucets/shower/kitchen. Run:
    python scripts/build-cube.py --validate      # accessory-only, check vs targets
    python scripts/build-cube.py                 # full cube -> src/data/cube.json
"""
import sys, os, json, datetime
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ACC = os.path.join(ROOT, "homedepot-reviews-bathroom-hardware-20260312 1.xlsx")
KB = os.path.join(ROOT, "homedepot-reviews-kitchen-bath-20260312 (1).xlsx")
OUT = os.path.join(ROOT, "src", "data", "cube.json")

# Column indices (0-based) in the source sheets.
C_CATEGORY, C_BRAND, C_DATE, C_STARS, C_SITE = 3, 4, 10, 13, 14
C_VERIFIED, C_RECOMMENDED, C_RESPONSEDATE = 17, 18, 28

SOURCES = ["Native (On-site)", "Brand Sites", "External Retail", "Social / Forums", "Other"]
SOCIAL = {"influenster.com"}
OTHER = {"felproductinfo.com", "lampsplus.com", "amerock corporation"}


def classify_source(site):
    """Return the source-channel index 0-4 for a ReviewFromSite value."""
    if site is None or str(site).strip() == "":
        return 0  # Native (on-site homedepot.com)
    s = str(site).strip().lower()
    if s == "homedepot.ca":
        return 2  # External Retail (Home Depot Canada)
    if s in SOCIAL:
        return 3
    if s in OTHER:
        return 4
    return 1  # Brand Sites (manufacturer domains)


def region_of(site):
    return "Canada" if (site is not None and str(site).strip().lower() == "homedepot.ca") else "United States"


def blank_cell():
    return {
        "n": 0, "ss": 0, "st": [0, 0, 0, 0, 0], "v": 0, "rec": 0, "resp": 0,
        "unv": 0, "uc": 0, "srcN": [0, 0, 0, 0, 0], "srcS": [0, 0, 0, 0, 0],
    }


def aggregate(files):
    cells = {}  # (brand, subcat, region, month) -> cell
    total = 0
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        for r in it:
            stars = r[C_STARS]
            if not isinstance(stars, int) or stars < 1 or stars > 5:
                continue
            brand = (str(r[C_BRAND]).strip() if r[C_BRAND] else "Unknown")
            subcat = (str(r[C_CATEGORY]).strip() if r[C_CATEGORY] else "Unknown")
            d = r[C_DATE]
            if not isinstance(d, datetime.datetime):
                continue
            month = "%04d-%02d" % (d.year, d.month)
            site = r[C_SITE]
            region = region_of(site)
            ch = classify_source(site)
            verified = str(r[C_VERIFIED]).strip() == "Yes"
            recommended = str(r[C_RECOMMENDED]).strip() == "Yes"
            responded = r[C_RESPONSEDATE] is not None

            key = (brand, subcat, region, month)
            c = cells.get(key)
            if c is None:
                c = blank_cell(); cells[key] = c
            c["n"] += 1
            c["ss"] += stars
            c["st"][stars - 1] += 1
            if verified:
                c["v"] += 1
            else:
                c["unv"] += 1
                if stars <= 2:
                    c["uc"] += 1
            if recommended:
                c["rec"] += 1
            if responded:
                c["resp"] += 1
            c["srcN"][ch] += 1
            c["srcS"][ch] += stars
            total += 1
        wb.close()
    return cells, total


def build(files, out_path, validate=False):
    cells, total = aggregate(files)

    brands = sorted({k[0] for k in cells})
    subcats = sorted({k[1] for k in cells})
    regions = ["United States", "Canada"]
    months = sorted({k[3] for k in cells})
    bi = {b: i for i, b in enumerate(brands)}
    si = {s: i for i, s in enumerate(subcats)}
    ri = {r: i for i, r in enumerate(regions)}
    mi = {m: i for i, m in enumerate(months)}

    rows = []
    for (brand, subcat, region, month), c in cells.items():
        rows.append([
            bi[brand], si[subcat], ri[region], mi[month],
            c["n"], c["ss"], *c["st"], c["v"], c["rec"], c["resp"], c["unv"], c["uc"],
            *c["srcN"], *c["srcS"],
        ])

    cube = {
        "meta": {"total": total, "generated": "2026-03-12", "grain": ["brand", "subcategory", "region", "month"]},
        "dict": {
            "brands": brands, "subcategories": subcats, "regions": regions,
            "sources": SOURCES, "months": months,
            "retailer": "The Home Depot",
            "category": "Kitchen & Bath",
        },
        "schema": ["b", "s", "r", "m", "n", "ss", "st1", "st2", "st3", "st4", "st5",
                   "v", "rec", "resp", "unv", "uc",
                   "srcN0", "srcN1", "srcN2", "srcN3", "srcN4",
                   "srcS0", "srcS1", "srcS2", "srcS3", "srcS4"],
        "cells": rows,
    }

    if validate:
        report(cube, cells, total)
    else:
        with open(out_path, "w", encoding="utf-8") as fh:
            json.dump(cube, fh, separators=(",", ":"))
        report(cube, cells, total)
        print("wrote", out_path, "(%.1f MB, %d cells)" % (os.path.getsize(out_path) / 1048576, len(rows)))


def report(cube, cells, total):
    def agg(pred):
        a = blank_cell()
        for k, c in cells.items():
            if pred and not pred(k):
                continue
            a["n"] += c["n"]; a["ss"] += c["ss"]; a["v"] += c["v"]; a["rec"] += c["rec"]
            a["resp"] += c["resp"]; a["unv"] += c["unv"]; a["uc"] += c["uc"]
            for i in range(5):
                a["st"][i] += c["st"][i]; a["srcN"][i] += c["srcN"][i]; a["srcS"][i] += c["srcS"][i]
        return a

    def m(a):
        n = a["n"] or 1
        return dict(reviews=a["n"], avg=round(a["ss"] / n, 3),
                    nativeShare=round(a["srcN"][0] / n, 4), verified=round(a["v"] / n, 4),
                    recommend=round(a["rec"] / n, 4), response=round(a["resp"] / n, 4),
                    unvComplaint=round(a["uc"] / (a["unv"] or 1), 4), srcCounts=a["srcN"])

    print("total reviews:", total, "| cells:", len(cells),
          "| brands:", len(cube["dict"]["brands"]), "| subcats:", len(cube["dict"]["subcategories"]),
          "| months:", len(cube["dict"]["months"]))
    print("TOTAL:", json.dumps(m(agg(None))))
    print("MOEN: ", json.dumps(m(agg(lambda k: k[0] == "MOEN"))))
    reg = {}
    for k, c in cells.items():
        reg[k[2]] = reg.get(k[2], 0) + c["n"]
    print("region split:", reg)


if __name__ == "__main__":
    if "--validate" in sys.argv:
        print("=== VALIDATION: accessory file only (should match shipped cube) ===")
        build([ACC], OUT, validate=True)
    else:
        print("=== FULL BUILD: accessories + kitchen/bath ===")
        build([ACC, KB], OUT, validate=False)
